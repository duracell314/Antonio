from openpyxl import Workbook, load_workbook
# Apro il file
wb = load_workbook("Trading_statistics.xlsx")
# ws è il foglio attivo
ws = wb.active

"""
Stampa Valore
"""
print(ws)
print(ws["A5"].value)

"""
Scrivi valore ( il file deve essere chiuso).
"""
# ws["A5"].value = "intel"
# wb.save("Trading_statistics.xlsx")

"""
Accedere ad altri sheets
"""
print(wb.sheetnames)

wb.create_sheet("Test")
print(wb.sheetnames)