import openpyxl

# In excel devo indicare la casella. Che è composta da una riga (numero) ed una colonna (lettera).

wb = openpyxl.load_workbook("../Prova.xlsx")

ws = wb['Operazioni']

print(ws['A5'])
# Se vogliamo leggere il valore occorre il metodo .value.
print(ws['A5'].value)


# Ma se usassimo i numeri al posto delle lettere?
# Voglio leggere A6 -> riga 6, colonna 1.

print(ws.cell(6, 1).value)

# Voglio scriverci?
ws.cell(6, 1).value = "STOCAZZO"
wb.save("../Prova.xlsx")

# Vogliamo estrarre più celle.
range = ws['A2':'D5']

# Questa istruzione dà errore.
# range = ws['A2':'B3'].value
# Come fare se vogliamo vedere il valore? Perchè?
# Perchè il risultato di ws[...] è una tupla, che non ha il metodo .value
# dobbiamo estrare la tupla.
print()
for a, b, c, d in range:
    print(a.value, b.value, c.value, d.value)


