import openpyxl

wb = openpyxl.load_workbook("../Prova.xlsx")
ws = wb['Operazioni']

# Iterare sulla riga.

rows = ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=2)
# Ci darà una tupla per ogni riga.
cols = ws.iter_cols(min_row=2, max_row=7, min_col=1, max_col=2)
# Ci darà una tupla per ogni colonna.
# A noi conviene usare iter_rows perchè poi avremmo tuple troppo grandi, difficili da gestire
#rows = ws.iter_rows(1, 7, 1, 4)
# print(rows)

# for row in rows:
#     print(row)
# ma così non possiamo leggere il valore, perchè row è una tupla.
# Dobbiamo estrarre la tupla.
Operazioni = []
Apertura = []
for a, b in rows:
    Operazioni.append(a.value)
    Apertura.append(b.value)

print(Operazioni)
print(Apertura)

# Se non conosciamo a priori la dimensione, ma vogliamo prendere tutti i dati basta non specificare.
# Nel nostro caso sappiamo che:
# le righe sono infinite, ma si parte da 2 perchè la prima riga è l'intestazione.
# le colonne sono 4.
# Quindi useremo:
rows = ws.iter_rows(min_row=2, min_col=1, max_col=4)
print("-" * 40)
Operazioni = []
Apertura = []
Chiusura = []
Risultato = []
for azione, apertura, chiusura, ris in rows:
    Operazioni.append(azione.value)
    Apertura.append(apertura.value)
    Chiusura.append(chiusura.value)
    Risultato.append(ris.value)

print(Operazioni, ": ", Risultato)
