import openpyxl

# Apriamo il workbook
wb = openpyxl.load_workbook("../Prova.xlsx")

# quali sono gli sheet?
print(wb.sheetnames)

# Selezionare il tuo sheet
ws = wb['Statistica per azione']

print(ws)
# Ci possiamo importare un altro sheet.
ws1 = wb['Statistiche']

# Possiamo creare un altro script
wb.create_sheet("Sheet_Nuovo")

# Così scelgo la posizione dello sheet.
wb.create_sheet("Sheet_Nuovo", 0)
print(wb.sheetnames)

# Alla fine dobbiamo salvare
wb.save("../Prova.xlsx")