from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import NewOperations as no


def check_if_duplicated() -> bool:
    """
    This function checks if what we want to insert is already in.
    So far was not possible to check the date, because the data we read from excel is a smaller data than the data we
    create in python environment (it contains also the hours, minutes and the seconds),
    so the comparison will always ba False.
    Bugfix is not urgent.
    :return: True if already inserted.
    """
    result = True
    for i in range(ROW_OFFSET, new_row + ROW_OFFSET):
        # TODO: check how to improve the comparison of date.
        # We cannot check the data now, for this reason we skip position 2 and 3.
        for j in range(1, 5, 3):
            op = no.New_Operations[i - ROW_OFFSET]
            d = op[j - 1]
            # Convert the column number into the letter
            c = get_column_letter(j)
            dummy = ws.cell(i, j).value
            if dummy == d:
                pass
            else:
                # If just one Data is different we have no duplicated data.
                result = False
                break
    return result


# "ROW OFFSET" is the row we actually start to work in excel sheet.
# Since the first is kept for the heading we start from position two.
ROW_OFFSET = 2
wb = load_workbook("Trading_statistics.xlsx")

# Select the proper sheet, otherwise we are not sure which one we will write to.
ws = wb['Operazioni']


# Calculate necessary rows
new_row = len(no.New_Operations)


# check is data is already inserted.
duplicated = check_if_duplicated()
# Now we insert financial data in excel blank rows
force_duplication = "N"
if duplicated:
    print("Data seems already inserted, if you want to force insertion type: Y")
    force_duplication = input(": ")
if (duplicated is False) or (force_duplication == "Y"):
    # Insert blank rows.
    ws.insert_rows(2, new_row)
    # ws.delete_rows(2, new_row)
    for row in range(ROW_OFFSET, new_row + ROW_OFFSET):
        for col in range(1, 5):
            operation = no.New_Operations[row - 2]
            data = operation[col - 1]
            # Convert the column number into the letter
            char = get_column_letter(col)
            ws.cell(row, col).value = data
    print("Data inserted")
else:
    print("Data not inserted")

# Save the results
wb.save("Trading_statistics.xlsx")


