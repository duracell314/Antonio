from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# Apro il file
# wb = load_workbook("Trading_statistics.xlsx")
"""
wb = Workbook()
ws = wb.active
ws.title = "Data"

ws.append(["Più", "parole", "in", "una", "istruzione"])
ws.append(["Più", "parole", "in", "una", "istruzione"])
ws.append(["Più", "parole", "in", "una", "istruzione"])
ws.append(["Più", "parole", "in", "una", "istruzione"])
ws.append(["prova"])
"""

wb = load_workbook("Trading_statistics.xlsx")
ws = wb.active
for row in range(1, 100):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)

"""
Inserire e cancellare righe
"""
ws.insert_rows(2)
wb.save("Trading_statistics.xlsx")