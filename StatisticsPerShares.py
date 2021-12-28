from openpyxl import Workbook, load_workbook
import Parameters as prm
import functions as functions
from datetime import datetime

wb = load_workbook("etoro-account-statement.xlsx")

# We create new sheet were we will put our data.
# We don't want duplicates, then if the sheet already exists we remove it before creating a new one:
sheets = wb.sheetnames
for sheet in sheets:
    if sheet == "Statistica_per_azione":
        std = wb["Statistica_per_azione"]
        wb.remove(std)
wb.create_sheet(title="Statistica_per_azione")

# Select the proper sheets.
# 'wo' is the sheet were are saved the operations
# 'ws' is the sheet were we will save the Statistics per shares
wo = wb['Posizioni chiuse']
wa = wb['Statistica_per_azione']

# Initialize the sheet:
for row in wa['A2:G1000']:
    for cell in row:
        cell.value = None

# We declare 2 lists, to divide and store the information.
All_Operations = []
All_Results = []
Operations = []
Number_of_Trades = []
Average_Results = []
Results = []
# We are not interested in the column 2 and 3. We are intetested on the column 1 and 4.
rows = wo.iter_rows(min_row=prm.ROW_START_OPERATIONS, min_col=prm.COLUMN_START_OPERATIONS, max_col=prm.COLUMN_END_OPERATIONS)

# We iterate through the tuples of rows and we add the data to a new element of the list
for id, stocks, imp, unity, op, cl, lev, spread, res, t_op, t_cl, t_tp, t_sl, comm, copy, typ, isin, note in rows:
    if copy.value == "-":
    # The operation was not copied.
        if stocks.value != None and res.value != None:
            # Covert a string (given by etoro report) into a date time variable.
            # '%Y' is used for the year format in 4 digits: 2021.
            # '%y' is used for the year format in 2 digits: 21.
            date_time_op_obj = datetime.strptime(op.value, '%d/%m/%Y %H:%M:%S')
            date_time_cl_obj = datetime.strptime(cl.value, '%d/%m/%Y %H:%M:%S')
            operation_days = functions.days_between(date_time_cl_obj, date_time_op_obj)
            if operation_days <= prm.MAXIMUM_DAYS_OPERATION_LENGTH:
                All_Operations.append(stocks.value)
                All_Results.append(res.value)
        elif stocks.value != None or res.value != None:
            # we enter in this conditional block if we have a partially filled row (for example if we missed one cell).
            All_Operations.append(None)
            All_Results.append(None)
            print("Erroneus data at excel rows: {0}".format(len(Operations + 1)))
        else:
            # All the elements ot the rows are empty, no operation needed.
            pass
    else:
    # We enter the else beacuse the operation was copied.
        if prm.COPYTRADER_ENABLE == False:
        # No operation needed, we don't want to add Copytrading in out statistics.
            pass
        else:
        # We add copy trading operations.
            if stocks.value != None and res.value != None:
                All_Operations.append(stocks.value)
                All_Results.append(res.value)
            elif stocks.value != None or res.value != None:
                # we enter in this conditional block if we have a partially filled row (for example if we missed one cell).
                All_Operations.append(None)
                All_Results.append(None)
                print("Erroneus data at excel rows: {0}".format(len(Operations + 1)))
            else:
                # All the elements ot the rows are empty, no operation needed.
                pass

# Now we create a list with only the stock we have traded and we sort it (otherwise we will find a different order everytime.
# We also initialize Number_of_Trades and Results list.
for stocks in All_Operations:
    if stocks not in Operations:
        Operations.append(stocks)
        Number_of_Trades.append(0)
        Results.append(0)
        Average_Results.append(0)
    else:
        pass
Operations.sort()

for index, stock in enumerate(Operations):
    for all_index, all_stock in enumerate(All_Operations):
        if stock == all_stock:
            Number_of_Trades[index] += 1
            Results[index] += All_Results[all_index]
            # Average_Results[index] = Results[index]/Number_of_Trades[index]
            # Results.append(res)

for index, stock in enumerate(Operations):
    Average_Results[index] = Results[index]/Number_of_Trades[index]


# Now it is time to save data into the excel sheet: the header and then the statistics
wa.cell(prm.ROW_START_STAT_PER_SHARE - 1, prm.COLUMN_SHARES_STAT_PER_SHARE).value = "Share"
wa.cell(prm.ROW_START_STAT_PER_SHARE - 1, prm.COLUMN_NUM_OF_SHARES_STAT_PER_SHARE).value = "Number of shares"
wa.cell(prm.ROW_START_STAT_PER_SHARE - 1, prm.COLUMN_TOTAL_RES_STAT_PER_SHARE).value = "Total result"
wa.cell(prm.ROW_START_STAT_PER_SHARE - 1, prm.COLUMN_AVERAGE_RES_STAT_PER_SHARE).value = "Average result"
for index in range(0, len(Operations)):
    wa.cell(prm.ROW_START_STAT_PER_SHARE + index, prm.COLUMN_SHARES_STAT_PER_SHARE).value = Operations[index]
    wa.cell(prm.ROW_START_STAT_PER_SHARE + index, prm.COLUMN_NUM_OF_SHARES_STAT_PER_SHARE).value = Number_of_Trades[index]
    wa.cell(prm.ROW_START_STAT_PER_SHARE + index, prm.COLUMN_TOTAL_RES_STAT_PER_SHARE).value = str(round(Results[index], 2)) + prm.CURRENCY
    wa.cell(prm.ROW_START_STAT_PER_SHARE + index, prm.COLUMN_AVERAGE_RES_STAT_PER_SHARE).value = str(round(Average_Results[index], 2)) + prm.CURRENCY

# TODO: Calcolare la migliore azione e la peggiore azione tradata.
# Save the statistics.
wb.save("etoro-account-statement.xlsx")