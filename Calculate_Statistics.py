from openpyxl import Workbook, load_workbook
import Parameters as prm
from datetime import datetime
from openpyxl.worksheet.copier import WorksheetCopy
# import StatisticsPerShares

wb = load_workbook("etoro-account-statement.xlsx")
# We create new sheet were we will put our data.
# We don't want duplicates, then if the sheet already exists we remove it before create a new one:
sheets = wb.sheetnames
for sheet in sheets:
    if sheet == "Statistiche":
        std = wb["Statistiche"]
        wb.remove(std)
wb.create_sheet(title="Statistiche")
# Select the proper sheets.


# 'wo' is the sheet were are saved the operations
# 'ws' is the sheet were we will save the Statistics
wo = wb['Posizioni chiuse']
ws = wb['Statistiche']
# Then we iterate through the rows of the operations done.
rows = wo.iter_rows(min_row=prm.ROW_START_OPERATIONS, min_col=prm.COLUMN_START_OPERATIONS, max_col=prm.COLUMN_END_OPERATIONS)

# We declare 4 lists, to divide and store the information.
Operations = []
Openings = []
Closings = []
Results = []
Leverages = []
Copied = []
Types = []
Notes = []

# We iterate through the tuples of rows and we add the data to a new element of the list
for id, stocks, imp, unity, open, close, lev, spread, res, t_op, t_cl, t_tp, t_sl, comm, copy, typ, isin, note in rows:
    # Check if the row is comletely filled.
    if stocks.value != None and open.value != None and close.value != None and res.value != None:
        if copy.value == "-":
            # The operation was not copied.
            Operations.append(stocks.value)
            # Covert a string (given by etoro report) into a date time variable.
            # '%Y' is used for the year format in 4 digits: 2021.
            # '%y' is used for the year format in 2 digits: 21.
            date_time_op_obj = datetime.strptime(open.value, '%d/%m/%Y %H:%M:%S')
            date_time_cl_obj = datetime.strptime(close.value, '%d/%m/%Y %H:%M:%S')
            Openings.append(date_time_op_obj)
            Closings.append(date_time_cl_obj)
            Results.append(res.value)
            Leverages.append(lev.value)
            Copied.append(copy.value)
            Types.append(typ.value)
            Notes.append(note.value)
        else:
        # We enter the else beacuse the operation was copied.
            if prm.COPYTRADER_ENABLE == False:
                # No operation needed, we don't want to add Copytrading in out statistics.
                pass
            else:
                # We add copy trading operations.
                Operations.append(stocks.value)
                # Covert a string (given by etoro report) into a date time variable.
                # '%Y' is used for the year format in 4 digits: 2021.
                # '%y' is used for the year format in 2 digits: 21.
                date_time_op_obj = datetime.strptime(open.value, '%d/%m/%Y %H:%M:%S')
                date_time_cl_obj = datetime.strptime(close.value, '%d/%m/%Y %H:%M:%S')
                Openings.append(date_time_op_obj)
                Closings.append(date_time_cl_obj)
                Results.append(res.value)
                Leverages.append(lev.value)
                Copied.append(copy.value)
                Types.append(typ.value)
                Notes.append(note.value)
    elif stocks.value != None or open.value != None or close.value != None or res.value != None:
        # we enter in this conditional block if we have a partially filled row (for example if we missed one cell).
        Operations.append(None)
        Openings.append(None)
        Closings.append(None)
        Results.append(None)
        Leverages.append(None)
        Copied.append(None)
        Types.append(None)
        Notes.append(None)
        print("Erroneus data at excel rows: {0}".format(len(Operations + 1)))
    else:
        # All the elements ot the rows are empty, no operation needed.
        pass

# We check if all the rows have the same length:
if len(Operations) == len(Openings) and len(Operations) == len(Closings) and len(Operations) == len(Results) and\
    len(Operations) == len(Leverages) and len(Operations) == len(Copied) and len(Operations) == len(Types) and\
    len(Operations) == len(Notes):
    # All it should be, no need to perform operations or raising any warning.
    pass
else:
    print("WARNING, not all the rows have the same length! Please check 'Operazioni' sheet in excel file")

# Timing data, will be necessary to calculate the statistics in a certain period of time.
last_day = max(Openings).date()
first_day = min(Openings).date()

# TODO: implementare i controlli sulle date. la data di fine deve essere minore della data di inizio ecc

total_operations = len(Operations)

# We calculate the number of operation with gain.
# This rule will be the same for all the variable metrics.
# _l suffix means last day
# _s suffix means selected day
# _p suffix means selected period

total_operations = len(Operations)
total_operations_l = 0
total_operations_s = 0
total_operations_p = 0

gain_operations = 0
gain_operations_l = 0
gain_operations_s = 0
gain_operations_p = 0
win = 0
win_l = 0
win_s = 0
win_p = 0
lose = 0
lose_l = 0
lose_s = 0
lose_p = 0
lose_operations = 0
lose_operations_l = 0
lose_operations_s = 0
lose_operations_p = 0

# General statistics
max_win = 0
max_lose = 0
earn_per_trade = 0
# TODO: implementare calcolo giorno migliore e giorno peggiore.
best_day = 0
worst_day = 0

for index, result in enumerate(Results):
    if result >= 0.0:
        gain_operations += 1
        if result > max_win:
            max_win = result
        win += result
        if Openings[index].date() == last_day:
            total_operations_l += 1
            gain_operations_l += 1
            win_l += result
        if Openings[index].date() == prm.selected_day:
            total_operations_s += 1
            gain_operations_s += 1
            win_s += result
        if Openings[index].date() >= prm.start_day and Openings[index].date() <= prm.end_day:
            total_operations_p += 1
            gain_operations_p += 1
            win_p += result
    else:
        # We calculate the number of operation with loses.
        lose_operations += 1
        lose += result
        if result < max_lose:
            max_lose = result
        if Openings[index].date() == last_day:
            total_operations_l += 1
            lose_operations_l += 1
            lose_l += result
        if Openings[index].date() == prm.selected_day:
            total_operations_s += 1
            lose_operations_s += 1
            lose_s += result
        if Openings[index].date() >= prm.start_day and Openings[index].date() <= prm.end_day:
            total_operations_p += 1
            lose_operations_p += 1
            lose_p += result

# NOTE: lose is already negative number.
net_income = win + lose
net_income_l = win_l + lose_l
net_income_s = win_s + lose_s
net_income_p = win_p + lose_p

if total_operations != 0:
    Batting_Average = round((100 * gain_operations / total_operations), 2)
else:
    Batting_Average = None

if total_operations_l != 0:
    Batting_Average_l = round((100 * gain_operations_l / total_operations_l), 2)
else:
    Batting_Average_l = None

if total_operations_s != 0:
    Batting_Average_s = round((100 * gain_operations_s / total_operations_s), 2)
else:
    Batting_Average_s = None

if total_operations_p != 0:
    Batting_Average_p = round((100 * gain_operations_p / total_operations_p), 2)
else:
    Batting_Average_p = None

if gain_operations != 0:
    Average_Win = win / gain_operations
else:
    Average_Win = 0

if gain_operations_l != 0:
    Average_Win_l = win_l / gain_operations_l
else:
    Average_Win_l = 0

if gain_operations_s != 0:
    Average_Win_s = win_s / gain_operations_s
else:
    Average_Win_s = 0

if gain_operations_p != 0:
    Average_Win_p = win_p / gain_operations_p
else:
    Average_Win_p = 0

if lose_operations != 0:
    Average_Lose = lose / lose_operations
else:
    Average_Lose = 0

if lose_operations_l != 0:
    Average_Lose_l = lose_l / lose_operations_l
else:
    Average_Lose_l = 0

if lose_operations_s != 0:
    Average_Lose_s = lose_s / lose_operations_s
else:
    Average_Lose_s = 0

if lose_operations_p != 0:
    Average_Lose_p = lose_p / lose_operations_p
else:
    Average_Lose_p = 0

# N.B. Average_Lose is a negative number

if Average_Lose != 0:
    Win_loss = Average_Win / (-Average_Lose)
else:
    Win_loss = 0

if Average_Lose_l != 0:
    Win_loss_l = Average_Win / (-Average_Lose_l)
else:
    Win_loss_l = 0

if Average_Lose_s != 0:
    Win_loss_s = Average_Win / (-Average_Lose_s)
else:
    Win_loss_s = 0

if Average_Lose_p != 0:
    Win_loss_p = Average_Win / (-Average_Lose_p)
else:
    Win_loss_p = 0

if total_operations != 0:
    earn_per_trade = net_income / total_operations
else:
    earn_per_trade = 0

# Write all data in the sheets.
ws.cell(prm.ROW_START_STATISTICS + 0, prm.COLUMN_STATISTIC - 1).value = "Number of operations"
ws.cell(prm.ROW_START_STATISTICS + 0, prm.COLUMN_STATISTIC).value = total_operations
ws.cell(prm.ROW_START_STATISTICS + 1, prm.COLUMN_STATISTIC - 1).value = "Net income"
ws.cell(prm.ROW_START_STATISTICS + 1, prm.COLUMN_STATISTIC).value = str(round(net_income, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 2, prm.COLUMN_STATISTIC - 1).value = "Operations in gain"
ws.cell(prm.ROW_START_STATISTICS + 2, prm.COLUMN_STATISTIC).value = gain_operations
ws.cell(prm.ROW_START_STATISTICS + 3, prm.COLUMN_STATISTIC - 1).value = "Operations in loss"
ws.cell(prm.ROW_START_STATISTICS + 3, prm.COLUMN_STATISTIC).value = lose_operations
ws.cell(prm.ROW_START_STATISTICS + 4, prm.COLUMN_STATISTIC - 1).value = "Batting average"
ws.cell(prm.ROW_START_STATISTICS + 4, prm.COLUMN_STATISTIC).value = str(Batting_Average) + "%"
ws.cell(prm.ROW_START_STATISTICS + 5, prm.COLUMN_STATISTIC - 1).value = "Win"
ws.cell(prm.ROW_START_STATISTICS + 5, prm.COLUMN_STATISTIC).value = str(round(win, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 6, prm.COLUMN_STATISTIC - 1).value = "Lose"
ws.cell(prm.ROW_START_STATISTICS + 6, prm.COLUMN_STATISTIC).value = str(round(lose, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 7, prm.COLUMN_STATISTIC - 1).value = "Average win"
ws.cell(prm.ROW_START_STATISTICS + 7, prm.COLUMN_STATISTIC).value = str(round(Average_Win, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 8, prm.COLUMN_STATISTIC - 1).value = "Average loss"
ws.cell(prm.ROW_START_STATISTICS + 8, prm.COLUMN_STATISTIC).value = str(round(Average_Lose, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 9, prm.COLUMN_STATISTIC - 1).value = "Win/Loss"
ws.cell(prm.ROW_START_STATISTICS + 9, prm.COLUMN_STATISTIC).value = str(round(Win_loss, 2)) + prm.CURRENCY

# Last day statistics
ws.cell(prm.ROW_START_STATISTICS - 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Last Day"
ws.cell(prm.ROW_START_STATISTICS - 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = last_day
ws.cell(prm.ROW_START_STATISTICS + 0 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Number of operations"
ws.cell(prm.ROW_START_STATISTICS + 0 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = total_operations_l
ws.cell(prm.ROW_START_STATISTICS + 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Net income"
ws.cell(prm.ROW_START_STATISTICS + 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(net_income_l, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 2 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Operations in gain"
ws.cell(prm.ROW_START_STATISTICS + 2 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = gain_operations_l
ws.cell(prm.ROW_START_STATISTICS + 3 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Operations in loss"
ws.cell(prm.ROW_START_STATISTICS + 3 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = lose_operations_l
ws.cell(prm.ROW_START_STATISTICS + 4 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Batting average"
ws.cell(prm.ROW_START_STATISTICS + 4 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(Batting_Average_l) + "%"
ws.cell(prm.ROW_START_STATISTICS + 5 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Win"
ws.cell(prm.ROW_START_STATISTICS + 5 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(win_l, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 6 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Lose"
ws.cell(prm.ROW_START_STATISTICS + 6 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(lose_l, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 7 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Average win"
ws.cell(prm.ROW_START_STATISTICS + 7 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(Average_Win_l, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 8 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Average loss"
ws.cell(prm.ROW_START_STATISTICS + 8 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(Average_Lose_l, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 9 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Win/Loss"
ws.cell(prm.ROW_START_STATISTICS + 9 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(Win_loss_l, 2)) + prm.CURRENCY

# Selected day statistics
ws.cell(prm.ROW_START_STATISTICS - 1 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Selected Day"
ws.cell(prm.ROW_START_STATISTICS - 1 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = prm.selected_day
ws.cell(prm.ROW_START_STATISTICS + 0 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Number of operations"
ws.cell(prm.ROW_START_STATISTICS + 0 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = total_operations_s
ws.cell(prm.ROW_START_STATISTICS + 1 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Net income"
ws.cell(prm.ROW_START_STATISTICS + 1 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(net_income_s, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 2 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Operations in gain"
ws.cell(prm.ROW_START_STATISTICS + 2 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = gain_operations_s
ws.cell(prm.ROW_START_STATISTICS + 3 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Operations in loss"
ws.cell(prm.ROW_START_STATISTICS + 3 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = lose_operations_s
ws.cell(prm.ROW_START_STATISTICS + 4 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Batting average"
ws.cell(prm.ROW_START_STATISTICS + 4 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(Batting_Average_s) + "%"
ws.cell(prm.ROW_START_STATISTICS + 5 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Win"
ws.cell(prm.ROW_START_STATISTICS + 5 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(win_s, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 6 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Lose"
ws.cell(prm.ROW_START_STATISTICS + 6 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(lose_s, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 7 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Average win"
ws.cell(prm.ROW_START_STATISTICS + 7 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(Average_Win_s, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 8 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Average loss"
ws.cell(prm.ROW_START_STATISTICS + 8 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(Average_Lose_s, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 9 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC - 1).value = "Win/Loss"
ws.cell(prm.ROW_START_STATISTICS + 9 + 2 * prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC).value = str(round(Win_loss_s, 2)) + prm.CURRENCY

# Selected period statistics
ws.cell(prm.ROW_START_STATISTICS - 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Selected period"
ws.cell(prm.ROW_START_STATISTICS - 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(prm.start_day) + " - " + str(prm.end_day)
ws.cell(prm.ROW_START_STATISTICS + 0 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Number of operations"
ws.cell(prm.ROW_START_STATISTICS + 0 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = total_operations_p
ws.cell(prm.ROW_START_STATISTICS + 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Net income"
ws.cell(prm.ROW_START_STATISTICS + 1 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    round(net_income_p, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 2 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Operations in gain"
ws.cell(prm.ROW_START_STATISTICS + 2 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = gain_operations_p
ws.cell(prm.ROW_START_STATISTICS + 3 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Operations in loss"
ws.cell(prm.ROW_START_STATISTICS + 3 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = lose_operations_p
ws.cell(prm.ROW_START_STATISTICS + 4 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Batting average"
ws.cell(prm.ROW_START_STATISTICS + 4 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    Batting_Average_p) + "%"
ws.cell(prm.ROW_START_STATISTICS + 5 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Win"
ws.cell(prm.ROW_START_STATISTICS + 5 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    round(win_p, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 6 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Lose"
ws.cell(prm.ROW_START_STATISTICS + 6 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    round(lose_p, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 7 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Average win"
ws.cell(prm.ROW_START_STATISTICS + 7 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    round(Average_Win_p, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 8 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Average loss"
ws.cell(prm.ROW_START_STATISTICS + 8 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    round(Average_Lose_p, 2)) + prm.CURRENCY
ws.cell(prm.ROW_START_STATISTICS + 9 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS - 1).value = "Win/Loss"
ws.cell(prm.ROW_START_STATISTICS + 9 + prm.ROW_SHIFT_STATISTICS, prm.COLUMN_STATISTIC + prm.COLUMN_SHIFT_STATISTICS).value = str(
    round(Win_loss_p, 2)) + prm.CURRENCY

# General statistics
ws["N4"].value = "Max win"
ws["O4"].value = str(round(max_win, 2)) + prm.CURRENCY
ws["N5"].value = "Max loss"
ws["O5"].value = str(round(max_lose, 2)) + prm.CURRENCY
ws["N6"].value = "First day"
ws["O6"].value = first_day
ws["N7"].value = "Last day"
ws["O7"].value = last_day
ws["N8"].value = "Earn per trade"
ws["O8"].value = str(round(earn_per_trade, 2)) + prm.CURRENCY
# ws["O9"].value = best_day
# ws["O10"].value = worst_day

# Save the statistics.
wb.save("etoro-account-statement.xlsx")
